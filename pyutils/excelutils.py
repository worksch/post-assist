from openpyxl import Workbook
from openpyxl import load_workbook
from enum import Enum

'''wb = Workbook()
ws = wb.active
ws['A1'] = 42
wb.save('/Users/weisf/Desktop/sample.xlsx')'''

'''title_course = ws['B2'].value       # 课序号
title_uuid = ws['C2'].value         # 序号
title_room = ws['D2'].value         # 班级
title_suid = ws['F2'].value         # 学号
title_gender = ws['H2'].value       # 性别
title_department = ws['I2'].value   # 院系
title_class = ws['J2'].value        # 班级
title_type = ws['K2'].value         # 学生类别
title_phone = ws['L2'].value        # 联系方式
title_email = ws['M2'].value        # 电了邮箱'''

class ExcelUtils():
    def open(self, xls):
        self.wb = load_workbook(xls)
        self.ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])

    def row_count(self):
        self.row_count = self.ws.max_row
        return self.row_count

    def col_count(self):
        self.col_count = self.ws.max_column
        return self.col_count

    def readColumn(self, row, col):
        return self.ws['{0}{1}'.format(chr(col), row)].value

    def read(self):
        row_count = row_count()
        col_count = col_count()
        for i in range(3, row_count):
            ws_course = self.ws['B{0}'.format(i)].value       # 课序号
            ws_uuid = self.ws['C{0}'.format(i)].value         # 序号
            ws_room = self.ws['D{0}'.format(i)].value         # 班级
            ws_suid = self.ws['E{0}'.format(i)].value         # 学号
            ws_name = self.ws['F{0}'.format(i)].value         # 姓名
            ws_gender = self.ws['H{0}'.format(i)].value       # 性别
            ws_department = self.ws['I{0}'.format(i)].value   # 院系
            ws_class = self.ws['J{0}'.format(i)].value        # 班级
            ws_type = self.ws['K{0}'.format(i)].value         # 学生类别
            ws_phone = self.ws['L{0}'.format(i)].value        # 联系方式
            ws_email = self.ws['M{0}'.format(i)].value        # 电了邮箱
            '''print ('{0},{1},{2},{3},{4},{5},{6},{7},{8},{9},{10}'
                    .format(ws_course, ws_uuid, ws_room, ws_suid, ws_gender, 
                        ws_department, ws_class, ws_type, ws_phone, ws_email))'''
            print('{0},{1},{2},{3},{4}'.format(ws_course, ws_uuid,ws_room, ws_suid, ws_name))

if __name__ == "__main__":
    utils = ExcelUtils()
    utils.open('/Users/weisf/Desktop/icenter/post-assistant/postassistant/static/files/excels/b88e9f5b063e43c38e162230d4470f6b1593884267759455.xlsx')
    
    print(utils.readColumn(3, 2))

class Column():
    A = 65
    B = 66
    C = 67
    D = 68
    E = 69
    F = 70
    G = 71
    H = 72
    I = 73
    J = 74
    K = 75
    L = 76
    M = 77
    N = 78
    O = 79
    P = 80
    Q = 81
    R = 82
    S = 83
    T = 84
    U = 85
    V = 86
    W = 87
    X = 88
    Y = 89
    Z = 90